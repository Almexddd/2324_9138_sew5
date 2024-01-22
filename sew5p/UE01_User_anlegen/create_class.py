import argparse
import logging
import random
from logging.handlers import RotatingFileHandler
from sys import stdout as out
from openpyxl import load_workbook

def generate_script(filepath) -> None:
    """
    generiert ein Script welches User für jede Klasse erstellt
    :param filepath: Pfad des Excelsheets mit den Klassen
    :return: ein script zum erstellen der user und eines zum löschen
    """
    with open('create_script.txt', 'w') as file:
        del_script = open('delete_script.txt', 'w')
        logins = open('logins.txt', 'w')
        wb = load_workbook(filepath, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for a in ws.iter_rows(min_row=2):
            if a[0].value is None:
                break
            classes = {
                "class": a[0].value,
                "RaumNr": a[1].value,
                "KV": a[2].value
            }
            pw = f"{classes['class']}{random.choice('!%&(),._-=^#')}{classes['RaumNr']}{random.choice('!%&(),._-=^#')}{classes['KV']}{random.choice('!%&(),._-=^#')}"
            file.write(f"useradd -d /home/klassen/{classes['class'].lower()} -c \"Klasse {classes['class']}\" -m -g {classes['class'].lower()} -G cdrom,plugdev,sambashare -s /bin/bash k{classes['class'].lower()} \n")
            file.write(f"echo '{pw}' | chpasswd\n")
            logger.debug(f"User k{classes['class'].lower()} created")
            del_script.write(f"userdel k{classes['class'].lower()} \n")
            logger.debug(f"delete script for user k{classes['class'].lower()} created")
            logins.write(f"k{classes['class'].lower()}: {pw} \n")
            logger.debug(f"login details for user k{classes['class'].lower()}")

parser = argparse.ArgumentParser()
parser.add_argument("filepath")
parser.add_argument("-v", "--verbosity")
parser.add_argument("-q", "--quiet")
args = parser.parse_args()
logging.basicConfig(level=logging.DEBUG if args.verbosity else logging.WARNING)
logger = logging.getLogger()
handler = RotatingFileHandler('logfile.log', maxBytes=10000, backupCount=5)
handler.setLevel(logging.DEBUG if args.verbosity else logging.WARNING)
logger.addHandler(handler)
stream_handler = logging.StreamHandler(out)
stream_handler.setLevel(logging.INFO)
logger.addHandler(stream_handler)
if args.filepath:
    generate_script(args.filepath)