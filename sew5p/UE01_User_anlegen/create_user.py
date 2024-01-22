"""
>>> sonderzeichen("di grillo")
'di_grillo'
>>> sonderzeichen("Gürbuz")
'Guerbuz'
>>> sonderzeichen("Rieß")
'Riess'
"""
import doctest
import argparse
import logging
import random
from unicodedata import normalize, category
from logging.handlers import RotatingFileHandler
from sys import stdout as out
from openpyxl import load_workbook
from openpyxl.workbook import Workbook


def sonderzeichen(s: str):
    """
    Ersetzt Umlaute, ß und Leerzeichen
    :param s: String in dem Sonderzeichen ersetzt werden sollen
    :return: Abgeänderter String
    """
    umlaute = {
        'ä': 'ae',
        'ö': 'oe',
        'ü': 'ue',
        'Ä': 'Ae',
        'Ö': 'Oe',
        'Ü': 'Ue',
        'ß': 'ss',
        ' ': '_'
    }
    for a, b in umlaute.items():
        s = s.replace(a, b)
    return s


def generate_script(filepath):
    """
    generiert das Script zum erstellen der User
    :param filepath: Pfad des Excelsheets, aus dem die User erstellt werden sollen
    :return: ein Script zum anlegen der User und eines zum löschen
    """
    with open('create_script.txt', 'w') as file:
        wb = load_workbook(filepath)
        ws = wb[wb.sheetnames[0]]
        cnt = 1
        user_cnt = dict()
        out = Workbook()
        ws_out =  out.active
        for user in ws.iter_rows(min_row=2):
            if user[0].value is None:
                break
            data = {
                "firstname": user[0].value,
                "lastname": user[1].value,
                "group": user[2].value,
                "class": user[3].value
            }
            data['firstname'] = sonderzeichen(data['firstname'])
            data['lastname'] = sonderzeichen(data['lastname'])
            data['firstname'] = ''.join(a for a in normalize('NFKD', data['firstname']) if category(a) != 'Mn')
            data['lastname'] = ''.join(a for a in normalize('NFKD', data['lastname']) if category(a) != 'Mn')
            pw = ''.join([random.choice('abcdefghijklmnopqrstuvwxyz!%&(),._-^#') for _ in range(20)])
            if data['lastname'] not in user_cnt:
                if data['group'] == 'teacher':
                    file.write(f"useradd -d /home/lehrer/{data['lastname']} -c \"{data['firstname']} {data['lastname']}\" -m -g {data['group']} -G cdrom,plugdev,sambashare -s /bin/bash {data['lastname'].lower()} \n")
                else:
                    file.write(f"useradd -d /home/klassen/k{data['class'].lower()}/{data['lastname']} -c \"{data['firstname']} {data['lastname']}\" -m -g {data['group']} -G cdrom,plugdev,sambashare -s /bin/bash {data['lastname'].lower()} \n")
                ws_out[f'A{cnt}'] = f'{data["lastname"].lower()}'
                ws_out[f'B{cnt}'] = f'{pw}'
                logger.debug(f'User {data["lastname"].lower()} added')
                user_cnt[data['lastname']] = 1

            else:
                if data['group'] == 'teacher':
                    file.write(f"useradd -d /home/lehrer/{data['lastname']} -c \"{data['firstname']} {data['lastname']}\" -m -g {data['group']} -G cdrom,plugdev,sambashare -s /bin/bash {data['lastname'].lower()}{user_cnt[data['lastname']]} \n")
                else:
                    file.write(f"useradd -d /home/klassen/k{data['class'].lower()}/{data['lastname']} -c \"{data['firstname']} {data['lastname']}\" -m -g {data['group']} -G cdrom,plugdev,sambashare -s /bin/bash {data['lastname'].lower()}{user_cnt[data['lastname']]} \n")
                ws_out[f'A{cnt}'] = f'{data["lastname"].lower()}{user_cnt[data["lastname"]]}'
                ws_out[f'B{cnt}'] = f'{pw}'
                logger.debug(f'User {data["lastname"].lower()}{user_cnt[data["lastname"]]} added')
                user_cnt[data['lastname']] += 1
            file.write(f"echo '{pw}' | chpasswd\n")
            cnt += 1
        out.save("Schüler_Passwörter.xlsx")

parser = argparse.ArgumentParser()
parser.add_argument("filepath")
parser.add_argument("-v", "--verbosity", help="increase output verbosity", action="store_true")
parser.add_argument("-q", "--quite")
args = parser.parse_args()

logging.basicConfig(level=logging.DEBUG if args.verbosity else logging.WARNING)
logger = logging.getLogger()
handler = RotatingFileHandler('logfile.log', maxBytes=10000, backupCount=5)
handler.setLevel(logging.DEBUG if args.verbosity else logging.WARNING)
logger.addHandler(handler)
stream_handler = logging.StreamHandler(out)
stream_handler.setLevel(logging.INFO)
logger.addHandler(stream_handler)
if args.filepath:
    generate_script(args.filepath)
    doctest.testmod()